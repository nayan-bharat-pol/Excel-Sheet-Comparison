import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("🔍 Excel File Comparison Tool (Dual Highlight + Excel + CSV Summary)")

# File uploaders
file1 = st.file_uploader("Upload First Excel File", type=["xls", "xlsx"])
file2 = st.file_uploader("Upload Second Excel File", type=["xls", "xlsx"])

if file1 and file2:
    # Read Excel files
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Normalize column names
    df1.columns = df1.columns.str.strip().str.replace(" ", "")
    df2.columns = df2.columns.str.strip().str.replace(" ", "")

    st.write("### File 1 Preview")
    st.dataframe(df1.head())
    st.write("### File 2 Preview")
    st.dataframe(df2.head())

    # Align both
    df1, df2 = df1.align(df2, join="outer", axis=1)
    df1 = df1.fillna("")
    df2 = df2.fillna("")

    # Compare
    diff_mask = df1 != df2
    total_changes = diff_mask.sum().sum()

    # Summary DataFrame for Streamlit + CSV
    diff_summary = []
    for r in df1.index:
        for c in df1.columns:
            if df1.at[r, c] != df2.at[r, c]:
                diff_summary.append({
                    "Row": r + 2,  # Excel-style (header offset)
                    "Column": c,
                    "File1_Value": df1.at[r, c],
                    "File2_Value": df2.at[r, c]
                })
    diff_summary_df = pd.DataFrame(diff_summary)

    st.write("### 🔸 Text Summary of Differences")
    st.dataframe(diff_summary_df)

    # --- Highlight differences in Excel ---
    def create_highlighted_workbook(df1, df2):
        """Create workbook with highlights and summary"""
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df1.to_excel(writer, sheet_name="File1", index=False)
            df2.to_excel(writer, sheet_name="File2", index=False)
        output.seek(0)

        wb = load_workbook(output)
        ws1 = wb["File1"]
        ws2 = wb["File2"]

        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for r in range(df1.shape[0]):
            for c, col in enumerate(df1.columns, 1):
                if df1.iat[r, c - 1] != df2.iat[r, c - 1]:
                    ws1.cell(row=r + 2, column=c).fill = yellow
                    ws2.cell(row=r + 2, column=c).fill = yellow

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "Comparison Summary"
        ws_summary["A2"] = f"Total cells changed: {total_changes}"
        ws_summary["A3"] = f"Rows in File1: {df1.shape[0]}"
        ws_summary["A4"] = f"Rows in File2: {df2.shape[0]}"
        ws_summary["A5"] = f"Columns compared: {len(df1.columns)}"

        return wb

    wb = create_highlighted_workbook(df1, df2)

    # Prepare Excel output
    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    # --- Download buttons ---
    st.download_button(
        label="📘 Download Highlighted Excel Report",
        data=excel_output,
        file_name="comparison_highlighted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if not diff_summary_df.empty:
        csv_output = diff_summary_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="📄 Download CSV Summary of Differences",
            data=csv_output,
            file_name="comparison_summary.csv",
            mime="text/csv"
        )

    # Summary Display
    st.success(f"✅ Comparison completed — {total_changes} total differences found across {len(df1.columns)} columns.")
    st.info("🟡 Yellow highlights = cells that differ\n📘 'Summary' sheet = overall stats\n📄 CSV file = full difference list")